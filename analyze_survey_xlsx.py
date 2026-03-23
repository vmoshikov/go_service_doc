#!/usr/bin/env python3
"""
Анализ Excel-файла с результатами опроса.
Определяет тип страницы по структуре строк:
- "N. Название" — название страницы
- Следующая строка "Вариант ответа" — страница с одним вопросом и множеством вариантов
- Следующая строка пустая — матрица (инструменты по строкам, варианты по колонкам)
- Следующая строка: вопрос + "Вариант ответа" — страница с множеством вопросов (подпункты)

Фильтрация (опционально, без --no-filter):
- Страницы по инструменту (напр. "13. GigaChat … Работа с контекстом"): из выборки
  исключаются респонденты, которые в матрице выбора инструментов ответили по этому
  инструменту "не знаком" или "знаю, но не использую" (вопрос задаётся через --tool-question).
- Матрица «опыт»: по-прежнему можно исключать "не использую"/"3" при подсчёте с опытом
  (см. _answer_should_exclude).
"""

import argparse
import csv
import json
import re
import sys
from collections import defaultdict
from pathlib import Path
from typing import Any, Dict, List, Optional, Set, Tuple, Union

try:
    import openpyxl
except ImportError:
    print("Требуется openpyxl: pip install openpyxl", file=sys.stderr)
    sys.exit(1)


# Паттерн: число. название (например "1. Ваша роль в команде")
PAGE_HEADER_PATTERN = re.compile(r"^\s*(\d+)\.\s+(.+)$")

# Страница «по инструменту»: "GigaChat .. …" или "13. GigaChat Работа с контекстом …"
DETAIL_PAGE_DOTDOT_PATTERN = re.compile(r"^([^.]+?)\s*\.\.\s*.+$")
DETAIL_PAGE_NUM_TOOL_PATTERN = re.compile(r"^\s*\d+\.\s*([^\s]+)")

# Ключевые фразы
VARIANT_OTVETA = "Вариант ответа"

# Для подсчёта «с опытом» в матрице (не для фильтра страниц по инструменту)
EXCLUDE_PHRASES = (
    "не знаю",
    "не использую",
    "не используется",
)
NO_EXPERIENCE_VALUES = ("3",)

# Первое слово после «N.» не считаем инструментом (обычные вопросы)
_TOOL_PAGE_FIRST_WORD_STOP = frozenset({
    "ваша", "ваш", "ваше", "какой", "какая", "какие", "какое", "оцените",
    "укажите", "выберите", "опишите", "насколько", "в", "как", "что", "есть",
    "при", "если", "для", "от", "до", "по", "на", "вопрос", "пройдите",
    "роль", "опыт", "уровень", "частота", "использование", "знакомство",
})

# Ответы в матрице инструментов → исключить респондента со страниц вопросов по этому инструменту
TOOL_DETAIL_EXCLUDE_PHRASES = (
    "не знаком",
    "знаю, но не использую",
    "знаю но не использую",
    "знаю, но не используется",
    "знаю но не используется",
)


def _get_cell_str(value) -> str:
    """Преобразует значение ячейки в строку."""
    if value is None:
        return ""
    return str(value).strip()


def _is_empty_row(row: tuple) -> bool:
    """Проверяет, пустая ли строка (все ячейки пустые или None)."""
    return all(_get_cell_str(c) == "" for c in row)


def _first_cell(row: tuple) -> str:
    """Первая ячейка строки."""
    return _get_cell_str(row[0]) if row else ""


def _has_variant_otveta(row: tuple) -> bool:
    """Есть ли в строке 'Вариант ответа'."""
    return any(VARIANT_OTVETA in _get_cell_str(c) for c in row)


def _is_page_header(cell: str) -> Tuple[bool, Optional[str], Optional[int]]:
    """
    Проверяет, является ли строка заголовком страницы (N. Название).
    Возвращает (True, название, номер) или (False, None, None).
    """
    m = PAGE_HEADER_PATTERN.match(cell)
    if m:
        return True, m.group(2).strip(), int(m.group(1))
    return False, None, None


def normalize_tool_key(name: str) -> str:
    """Ключ для сопоставления инструмента в матрице и в заголовке страницы."""
    return name.strip().lower()


def extract_tool_from_page_name(page_name: str) -> Optional[str]:
    """
    Имя инструмента для страниц вопросов по инструменту:
    "GigaChat .. …" или "13. GigaChat Работа с контекстом …"
    """
    s = page_name.strip()
    m = DETAIL_PAGE_DOTDOT_PATTERN.match(s)
    if m:
        return m.group(1).strip()
    m2 = DETAIL_PAGE_NUM_TOOL_PATTERN.match(s)
    if m2:
        w = m2.group(1).strip()
        if w.lower() in _TOOL_PAGE_FIRST_WORD_STOP:
            return None
        return w
    return None


def _answer_should_exclude(answer: str, exclude_rating_3: bool = True) -> bool:
    """Проверяет, исключает ли ответ респондента (нет опыта с инструментом)."""
    a = answer.lower().strip()
    if any(phrase in a for phrase in EXCLUDE_PHRASES):
        return True
    if exclude_rating_3 and a in NO_EXPERIENCE_VALUES:
        return True
    return False


def _answer_excludes_from_tool_detail(answer: str) -> bool:
    """Ответ в матрице инструментов → не показывать этого респондента на страницах по инструменту."""
    a = answer.lower().strip()
    return any(phrase in a for phrase in TOOL_DETAIL_EXCLUDE_PHRASES)


def get_main_questions(pages: List[dict]) -> List[Tuple[int, str]]:
    """Возвращает пронумерованный список основных вопросов (страниц)."""
    return [(p.get("num", p["row"]), p["name"]) for p in pages]


def load_survey_rows(filepath: Union[str, Path]) -> List[tuple]:
    """Загружает все строки листа."""
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    return rows


def build_respondent_exclusion(
    rows: List[tuple],
    pages: List[dict],
    tool_question_num: int = 9,
    total_cols: int = 0,
) -> Dict[str, Set[int]]:
    """
    По матрице выбора инструментов: для каждого инструмента — столбцы респондентов,
    ответивших «не знаком» / «знаю, но не использую» (исключать со страниц по инструменту).
    Ключи инструментов — normalize_tool_key (нижний регистр).
    """
    exclusion: Dict[str, Set[int]] = {}
    # Ищем строки вопроса tool_question_num: страницы с этим номером и следующие
    # (матрица: инструменты по строкам, ответы по столбцам)
    tool_page_rows: List[int] = []
    for p in pages:
        if p.get("num") == tool_question_num:
            tool_page_rows.append(p["row"] - 1)  # 0-based
            break

    if not tool_page_rows:
        return exclusion

    tc = total_cols or (len(rows[0]) if rows else 0)
    scan_cols = set(range(1, max(tc, 1)))

    start_row = tool_page_rows[0]
    # Следующие строки до следующей страницы — строки инструментов
    next_page_row = len(rows)
    for p in pages:
        if p["row"] - 1 > start_row:
            next_page_row = p["row"] - 1
            break

    # Пропускаем строку заголовков (пустая или "Вариант ответа")
    data_start = start_row + 1
    if data_start < len(rows) and (
        _is_empty_row(rows[data_start]) or _has_variant_otveta(rows[data_start])
    ):
        data_start += 1

    for r in range(data_start, next_page_row):
        if r >= len(rows):
            break
        row = rows[r]
        tool_name = _get_cell_str(row[0]) if row else ""
        if not tool_name or _is_page_header(tool_name)[0]:
            continue
        for col_idx in scan_cols:
            if col_idx >= len(row):
                continue
            val = _get_cell_str(row[col_idx])
            if val and _answer_excludes_from_tool_detail(val):
                nk = normalize_tool_key(tool_name)
                exclusion.setdefault(nk, set()).add(col_idx)

    return exclusion


def get_valid_respondent_columns(
    exclude_for_tool: Optional[str],
    exclusion_map: Dict[str, Set[int]],
    base_columns: Set[int],
) -> Set[int]:
    """
    Возвращает индексы столбцов респондентов, которые учитываются.
    base_columns — допустимые столбцы для данной страницы (напр. B–F для вопроса 9).
    """
    valid = set(base_columns)
    if not exclude_for_tool:
        return valid
    k = normalize_tool_key(exclude_for_tool)
    if k in exclusion_map:
        valid -= exclusion_map[k]
    return valid


def _get_page_block_bounds(
    rows: List[tuple],
    pages: List[dict],
    page_idx: int,
) -> Tuple[int, int]:
    """Возвращает (start_row, end_row) 0-based для блока страницы."""
    start = pages[page_idx]["row"] - 1
    end = len(rows)
    if page_idx + 1 < len(pages):
        end = pages[page_idx + 1]["row"] - 1
    return start, end


def detect_layout(rows: List[tuple]) -> str:
    """
    Определяет раскладку: 'columns' = респонденты в столбцах (col 1..N),
    'rows' = респонденты в строках (row 1..N).
    Эвристика: если в первой строке есть паттерн N. в нескольких ячейках — заголовки по колонкам.
    Если в первом столбце много N. — структура по строкам.
    """
    if not rows:
        return "columns"
    first_row = rows[0]
    first_col_headers = sum(
        1 for c in first_row if c and PAGE_HEADER_PATTERN.match(_get_cell_str(c))
    )
    first_col_values = sum(
        1 for r in rows[:50] if r and PAGE_HEADER_PATTERN.match(_get_cell_str(r[0]))
    )
    if first_col_values > first_col_headers:
        return "columns"  # N. в первом столбце — классическая структура
    return "rows"  # N. в первой строке — респонденты в строках


def extract_question_stats(
    rows: List[tuple],
    pages: List[dict],
    page_idx: int,
    valid_columns: Optional[Set[int]] = None,
    exclude_rating_3: bool = True,
) -> Dict[str, Any]:
    """
    Извлекает числовую статистику по ответам для одной страницы.
    valid_columns: индексы столбцов респондентов для учёта (None = все).
    """
    total_cols = len(rows[0]) if rows else 0
    p = pages[page_idx]
    start_row, end_row = _get_page_block_bounds(rows, pages, page_idx)
    page_type = p["type"]

    if valid_columns is None:
        valid_columns = set(range(1, total_cols))

    result = {
        "page_num": p["num"],
        "page_name": p["name"],
        "type": page_type,
        "tool": p.get("tool"),
        "total_respondents": len(valid_columns),
        "answers": defaultdict(int),
        "matrix": {},
        "matrix_n": {},  # по инструменту: кол-во респондентов с опытом (исключая "не использую")
        "subquestions": [],
    }

    if start_row >= len(rows):
        return dict(result)

    # Пропуск строки заголовка страницы
    data_start = start_row + 1
    if data_start >= end_row:
        return dict(result)

    header_row = rows[data_start]
    data_start += 1

    if page_type == "single_question":
        # Один вопрос: каждый респондент (колонка) даёт один ответ.
        # Сумма по колонкам: для каждой колонки берём первый непустой ответ в блоке.
        for col in valid_columns:
            val = None
            for r in range(data_start, end_row):
                if r >= len(rows):
                    break
                row = rows[r]
                if _has_variant_otveta(row) or _is_empty_row(row):
                    continue
                if col < len(row):
                    v = _get_cell_str(row[col])
                    if v and v != VARIANT_OTVETA and not _is_page_header(v)[0]:
                        val = v
                        break
            if val:
                result["answers"][val] += 1

    elif page_type == "matrix":
        # Строки = инструменты, столбцы = респонденты, ячейка = ответ.
        # Исключаем респондентов, ответивших "не использую" / "3" для данного инструмента —
        # считаем только тех, кто имеет опыт использования.
        if data_start < len(rows) and (
            _is_empty_row(rows[data_start]) or _has_variant_otveta(rows[data_start])
        ):
            data_start += 1
        for r in range(data_start, end_row):
            if r >= len(rows):
                break
            row = rows[r]
            row_label = _get_cell_str(row[0]) if row else ""
            if not row_label or _is_page_header(row_label)[0]:
                continue
            result["matrix"][row_label] = defaultdict(int)
            respondents_with_experience = 0
            for col in valid_columns:
                if col < len(row):
                    val = _get_cell_str(row[col])
                    if val:
                        if _answer_should_exclude(val, exclude_rating_3):
                            continue  # Исключаем: нет опыта с этим инструментом
                        result["matrix"][row_label][val] += 1
                        respondents_with_experience += 1
            result["matrix_n"][row_label] = respondents_with_experience

    elif page_type == "multiple_questions":
        # Строки = подвопросы (вопрос в col 0), столбцы = респонденты
        r = data_start
        while r < end_row and r < len(rows):
            row = rows[r]
            q_text = _get_cell_str(row[0]) if row else ""
            if not q_text or _is_page_header(q_text)[0]:
                r += 1
                continue
            sub_answers = defaultdict(int)
            for col in valid_columns:
                if col < len(row):
                    val = _get_cell_str(row[col])
                    if val:
                        sub_answers[val] += 1
            result["subquestions"].append({
                "question": q_text,
                "answers": dict(sub_answers),
            })
            r += 1

    result["answers"] = dict(result["answers"])
    result["matrix"] = {k: dict(v) for k, v in result["matrix"].items()}
    return result


def _get_page_column_bounds(rows: List[tuple]) -> List[Tuple[int, int, dict]]:
    """
    Для layout=rows: сканирует строку 0, находит заголовки N. и возвращает
    [(start_col, end_col, page_info), ...].
    """
    if not rows:
        return []
    header = rows[0]
    result = []
    i = 0
    while i < len(header):
        cell = _get_cell_str(header[i])
        is_h, name, num = _is_page_header(cell)
        if is_h and name:
            start = i
            end = i + 1
            j = i + 1
            while j < len(header):
                next_cell = _get_cell_str(header[j])
                if _is_page_header(next_cell)[0]:
                    end = j
                    break
                j += 1
            else:
                end = len(header)
            result.append((start, end, {"num": num, "name": name, "row": 1}))
            i = end
        else:
            i += 1
    return result


def extract_question_stats_rows_layout(
    rows: List[tuple],
    col_start: int,
    col_end: int,
    page_info: dict,
    valid_rows: Optional[Set[int]] = None,
) -> Dict[str, Any]:
    """
    Извлечение статистики когда респонденты в строках (rows 1..N).
    col_start, col_end — диапазон колонок для этого вопроса.
    """
    total_rows = len(rows) - 1 if rows else 0  # минус заголовок
    if valid_rows is None:
        valid_rows = set(range(1, len(rows)))
    result = {
        "page_num": page_info.get("num", 0),
        "page_name": page_info.get("name", ""),
        "type": "single_question",
        "tool": extract_tool_from_page_name(page_info.get("name", "")),
        "total_respondents": len(valid_rows),
        "answers": defaultdict(int),
        "matrix": {},
        "subquestions": [],
    }
    for r in valid_rows:
        if r >= len(rows):
            continue
        row = rows[r]
        for c in range(col_start, min(col_end, len(row))):
            val = _get_cell_str(row[c])
            if val:
                result["answers"][val] += 1
    result["answers"] = dict(result["answers"])
    return result


def extract_all_stats(
    rows: List[tuple],
    pages: List[dict],
    exclusion_map: Optional[Dict[str, Set[int]]] = None,
    tool_question_num: int = 9,
    layout: str = "columns",
    exclude_rating_3: bool = True,
) -> List[Dict[str, Any]]:
    """
    Извлекает статистику по всем страницам.
    Применяет фильтрацию по инструментам для страниц детализации.
    layout: 'columns' = респонденты в столбцах, 'rows' = респонденты в строках.
    """
    if layout == "rows":
        col_bounds = _get_page_column_bounds(rows)
        all_stats = []
        for start, end, page_info in col_bounds:
            stats = extract_question_stats_rows_layout(
                rows, start, end, page_info, None
            )
            all_stats.append(stats)
        return all_stats

    total_cols = len(rows[0]) if rows else 0
    all_stats = []

    for i, p in enumerate(pages):
        base_cols = set(range(1, total_cols))
        valid = None
        if exclusion_map and p.get("tool") and total_cols > 1:
            valid = get_valid_respondent_columns(
                p["tool"], exclusion_map, base_cols
            )
        stats = extract_question_stats(
            rows, pages, i, valid, exclude_rating_3=exclude_rating_3
        )
        all_stats.append(stats)

    return all_stats


def build_text_report(
    stats_list: List[Dict[str, Any]],
    validate: bool = False,
) -> str:
    """Формирует текстовый отчёт."""
    lines = []
    lines.append("=" * 80)
    lines.append("ОТЧЁТ ПО РЕЗУЛЬТАТАМ ОПРОСА")
    lines.append(f"Всего вопросов/страниц: {len(stats_list)}")
    total_resp = max((s["total_respondents"] for s in stats_list), default=0)
    lines.append(f"Макс. респондентов на странице: {total_resp}")
    lines.append("=" * 80)

    for s in stats_list:
        lines.append("")
        lines.append(f"{s['page_num']}. {s['page_name']}")
        if s.get("tool"):
            lines.append(f"   [Инструмент: {s['tool']}] | Респондентов: {s['total_respondents']}")
        else:
            lines.append(f"   Респондентов: {s['total_respondents']}")
        lines.append("-" * 60)

        if s["type"] == "single_question" and s["answers"]:
            total_ans = sum(s["answers"].values())
            if validate or (total_ans != s["total_respondents"] and total_ans > 0):
                lines.append(
                    f"   [Сумма ответов по колонкам: {total_ans}, респондентов: {s['total_respondents']}]"
                )
            for ans, cnt in sorted(s["answers"].items(), key=lambda x: -x[1]):
                pct = 100 * cnt / s["total_respondents"] if s["total_respondents"] else 0
                lines.append(f"   {ans}: {cnt} ({pct:.1f}%)")

        elif s["type"] == "matrix" and s["matrix"]:
            matrix_n = s.get("matrix_n", {})
            for row_label, answers in s["matrix"].items():
                n_with_experience = matrix_n.get(row_label, sum(answers.values()))
                excluded = s["total_respondents"] - n_with_experience
                lines.append(
                    f"   {row_label}: {n_with_experience} с опытом"
                    + (f" (исключено {excluded} без опыта)" if excluded else "")
                    + ":"
                )
                row_total = sum(answers.values())
                for ans, cnt in sorted(answers.items(), key=lambda x: -x[1]):
                    pct = 100 * cnt / row_total if row_total else 0
                    lines.append(f"      {ans}: {cnt} ({pct:.1f}%)")

        elif s["type"] == "multiple_questions" and s["subquestions"]:
            for sq in s["subquestions"]:
                sq_total = sum(sq["answers"].values())
                if validate:
                    lines.append(f"   {sq['question']} [сумма: {sq_total}]:")
                else:
                    lines.append(f"   {sq['question']}:")
                for ans, cnt in sorted(sq["answers"].items(), key=lambda x: -x[1]):
                    pct = 100 * cnt / sq_total if sq_total else 0
                    lines.append(f"      {ans}: {cnt} ({pct:.1f}%)")

    return "\n".join(lines)


def build_csv_report(stats_list: List[Dict[str, Any]]) -> str:
    """Формирует CSV-отчёт (плоская структура для импорта в Excel)."""
    import io
    buf = io.StringIO()
    w = csv.writer(buf, delimiter=";", quoting=csv.QUOTE_MINIMAL)
    w.writerow([
        "Номер", "Страница", "Тип", "Инструмент", "Вопрос/Строка",
        "С опытом (N)", "Вариант ответа", "Кол-во", "%"
    ])

    for s in stats_list:
        n = s["page_num"]
        name = s["page_name"]
        t = s["type"]
        tool = s.get("tool") or ""
        total = s["total_respondents"]

        if t == "single_question":
            for ans, cnt in s["answers"].items():
                pct = 100 * cnt / total if total else 0
                w.writerow([n, name, t, tool, "", "", ans, cnt, f"{pct:.1f}"])

        elif t == "matrix":
            matrix_n = s.get("matrix_n", {})
            for row_label, answers in s["matrix"].items():
                n_exp = matrix_n.get(row_label, sum(answers.values()))
                row_total = sum(answers.values())
                for ans, cnt in answers.items():
                    pct = 100 * cnt / row_total if row_total else 0
                    w.writerow([n, name, t, tool, row_label, n_exp, ans, cnt, f"{pct:.1f}"])

        elif t == "multiple_questions":
            for sq in s["subquestions"]:
                q = sq["question"]
                row_total = sum(sq["answers"].values())
                for ans, cnt in sq["answers"].items():
                    pct = 100 * cnt / row_total if row_total else 0
                    w.writerow([n, name, t, tool, q, "", ans, cnt, f"{pct:.1f}"])

    return buf.getvalue()


def _parse_pages_filter(spec: str) -> Set[int]:
    """Парсит '1,3,5-10' в множество номеров страниц."""
    result = set()
    for part in spec.split(","):
        part = part.strip()
        if "-" in part:
            a, b = part.split("-", 1)
            try:
                result.update(range(int(a.strip()), int(b.strip()) + 1))
            except ValueError:
                pass
        else:
            try:
                result.add(int(part))
            except ValueError:
                pass
    return result


def build_json_report(stats_list: List[Dict[str, Any]]) -> str:
    """Формирует JSON-отчёт."""
    # Конвертируем defaultdict в dict для JSON
    out = []
    for s in stats_list:
        out.append({
            "page_num": s["page_num"],
            "page_name": s["page_name"],
            "type": s["type"],
            "tool": s.get("tool"),
            "total_respondents": s["total_respondents"],
            "answers": s["answers"],
            "matrix": s["matrix"],
            "matrix_n": s.get("matrix_n", {}),
            "subquestions": s["subquestions"],
        })
    return json.dumps(out, ensure_ascii=False, indent=2)


def analyze_survey_xlsx(filepath: Union[str, Path]) -> List[dict]:
    """
    Анализирует Excel-файл и возвращает список страниц с их типами.

    Возвращает:
        [
            {"row": int, "name": str, "type": str, "description": str},
            ...
        ]
        type: "single_question" | "matrix" | "multiple_questions"
    """
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    result = []

    i = 0
    while i < len(rows):
        row = rows[i]
        first = _first_cell(row)

        is_header, page_name, page_num = _is_page_header(first)
        if not is_header or not page_name:
            i += 1
            continue

        next_row = rows[i + 1] if i + 1 < len(rows) else ()
        next_first = _first_cell(next_row)
        next_has_variant = _has_variant_otveta(next_row)

        page_type = "unknown"
        description = ""

        if not next_first or _is_empty_row(next_row):
            page_type = "matrix"
            description = "Вопрос в формате матрицы: по строкам — инструменты, по колонкам — варианты использования"
        elif next_has_variant and VARIANT_OTVETA in next_first:
            page_type = "single_question"
            description = "Страница с одним вопросом и множеством вариантов ответа"
        elif next_has_variant and next_first:
            page_type = "multiple_questions"
            description = "Страница с множеством вопросов (подпункты)"

        result.append({
            "row": i + 1,
            "num": page_num,
            "name": page_name,
            "type": page_type,
            "description": description,
            "tool": extract_tool_from_page_name(page_name),
        })
        i += 1

    return result


def main():
    parser = argparse.ArgumentParser(
        description="Анализ Excel-файла с результатами опроса. "
        "Вывод основных вопросов и фильтрация по ответам на вопрос об инструментах."
    )
    parser.add_argument(
        "file",
        nargs="?",
        default=None,
        help="Путь к xlsx-файлу (по умолчанию result_summary.xlsx или analyze_demo.xlsx)",
    )
    parser.add_argument(
        "--list-questions",
        "-q",
        action="store_true",
        help="Вывести пронумерованный список основных вопросов",
    )
    parser.add_argument(
        "--filter-tool",
        "-f",
        metavar="ИНСТРУМЕНТ",
        help="Применить фильтр: исключить респондентов, ответивших "
        "'не знаю/не использую' на этот инструмент (для страниц детализации)",
    )
    parser.add_argument(
        "--tool-question",
        "-t",
        type=int,
        default=9,
        metavar="N",
        help="Номер вопроса о выборе инструментов (по умолчанию 9)",
    )
    parser.add_argument(
        "--no-filter",
        action="store_true",
        help="Не применять фильтрацию (показать всех респондентов)",
    )
    parser.add_argument(
        "--no-filter-tool-detail",
        action="store_true",
        help="Не исключать по ответам в матрице инструментов "
        "(«не знаком» / «знаю, но не использую») на страницах по инструменту",
    )
    parser.add_argument(
        "--report",
        "-r",
        action="store_true",
        help="Построить расширенный отчёт с числовыми значениями по каждому типу ответов",
    )
    parser.add_argument(
        "--format",
        choices=["text", "csv", "json"],
        default="text",
        help="Формат отчёта (по умолчанию text)",
    )
    parser.add_argument(
        "--output",
        "-o",
        metavar="FILE",
        help="Сохранить отчёт в файл",
    )
    parser.add_argument(
        "--pages",
        "-p",
        metavar="N,M-K",
        help="Только указанные страницы (например: 1,3,5-10)",
    )
    parser.add_argument(
        "--layout",
        choices=["columns", "rows", "auto"],
        default="auto",
        help="Респонденты в столбцах (columns) или в строках (rows). "
        "columns: колонки 1..N = респонденты. rows: строки 1..N = респонденты.",
    )
    parser.add_argument(
        "--validate",
        action="store_true",
        help="Проверка: выводить сумму ответов vs респондентов для каждого вопроса",
    )
    parser.add_argument(
        "--no-exclude-3",
        action="store_true",
        help="Не исключать оценку 3 (по умолчанию 3 = нет опыта, исключается)",
    )
    args = parser.parse_args()

    filepath = Path(args.file) if args.file else Path("result_summary.xlsx")
    if not filepath.exists():
        filepath = Path("analyze_demo.xlsx")
    if not filepath.exists():
        print(f"Файл не найден: {filepath}", file=sys.stderr)
        sys.exit(1)

    pages = analyze_survey_xlsx(filepath)
    rows = load_survey_rows(filepath)
    total_cols = len(rows[0]) if rows else 0
    exclusion_map = {}
    if not args.no_filter and not args.no_filter_tool_detail:
        exclusion_map = build_respondent_exclusion(
            rows, pages, args.tool_question, total_cols
        )

    if args.report:
        layout = args.layout
        if layout == "auto":
            layout = detect_layout(rows)
        stats_list = extract_all_stats(
            rows,
            pages,
            exclusion_map,
            args.tool_question,
            layout,
            exclude_rating_3=not args.no_exclude_3,
        )
        if args.pages:
            page_nums = _parse_pages_filter(args.pages)
            stats_list = [s for s in stats_list if s["page_num"] in page_nums]
        if args.format == "text":
            report_text = build_text_report(stats_list, validate=args.validate)
        elif args.format == "csv":
            report_text = build_csv_report(stats_list)
        else:
            report_text = build_json_report(stats_list)
        if args.output:
            Path(args.output).write_text(report_text, encoding="utf-8")
            print(f"Отчёт сохранён: {args.output}", file=sys.stderr)
        else:
            print(report_text)
        return

    if args.list_questions:
        print("Основные вопросы:\n")
        for num, name in get_main_questions(pages):
            tool = extract_tool_from_page_name(name)
            suffix = f" [инструмент: {tool}]" if tool else ""
            print(f"  {num}. {name}{suffix}")
        print()

    print(f"Файл: {filepath}\n")
    print(f"{'Строка':<6} {'Тип':<20} {'Название страницы'}")
    print("-" * 80)

    for p in pages:
        line = f"{p['row']:<6} {p['type']:<20} {p['name']}"
        filter_info = ""
        if not args.no_filter and not args.no_filter_tool_detail and p.get("tool") and total_cols > 1:
            _base = set(range(1, total_cols))
            valid = get_valid_respondent_columns(p["tool"], exclusion_map, _base)
            excluded = len(_base) - len(valid)
            if excluded > 0:
                filter_info = f" (респондентов с фильтром: {len(valid)}/{len(_base)}, исключено: {excluded})"
        elif args.filter_tool and total_cols > 1:
            _base = set(range(1, total_cols))
            valid = get_valid_respondent_columns(
                args.filter_tool, exclusion_map, _base
            )
            excluded = len(_base) - len(valid)
            if p.get("tool") == args.filter_tool and excluded > 0:
                filter_info = f" (учтено: {len(valid)}, исключено: {excluded})"
        print(line + filter_info)
        if p["description"]:
            print(f"       └─ {p['description']}")

    if exclusion_map and not args.no_filter and not args.no_filter_tool_detail:
        print("\n--- Фильтрация страниц по инструменту ---")
        print("Исключаются респонденты с ответами «не знаком» / «знаю, но не использую»")
        print("в матрице инструментов (--tool-question).")
        for tool, cols in sorted(exclusion_map.items()):
            print(f"  {tool}: исключено {len(cols)} респондентов")


if __name__ == "__main__":
    main()
